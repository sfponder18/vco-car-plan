import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=14, color='FFFFFF')
subheader_font = Font(bold=True, size=11, color='FFFFFF')
label_font = Font(bold=True, size=10)
normal_font = Font(size=10)

dark_fill = PatternFill('solid', fgColor='1F2937')
blue_fill = PatternFill('solid', fgColor='2563EB')
green_fill = PatternFill('solid', fgColor='00B050')
red_fill = PatternFill('solid', fgColor='FF0000')
yellow_fill = PatternFill('solid', fgColor='FFFF00')
light_gray = PatternFill('solid', fgColor='F3F4F6')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

# Hotel color map
hotel_fills = {
    'Flexbase': yellow_fill,
    'Kysthotell': green_fill,
    'Fosen Fjord': red_fill,
}

# Authorized fuelers
fuelers = [
    'Merritt, Dustin', 'Williams, Sean', 'Olmschenk, Cole',
    'Cunningham-Wray, Kassidy', 'Hanlin, Stephanie', 'Moroz, Christopher',
    'Quimby, James', 'Ponder, Stephen', 'Kolmer, Benjamin',
    'Ortega, Soleil', 'Evans, Noah', 'Moore, Jaden'
]

# Full personnel list
personnel = [
    ('Merritt, Dustin M.', 'Shaggi', 'Pilots', 'Flexbase'),
    ('Cole, Charles J.', '', 'Pilots', 'Flexbase'),
    ('Olmschenk, Cole E.', '', 'Pilots', 'Flexbase'),
    ('Gould, Joseph A.', '', 'Pilots', 'Flexbase'),
    ('Thompson, Kegan M.', '', 'Pilots', 'Flexbase'),
    ('Cartin, Daniel M.', '', 'Pilots', 'Flexbase'),
    ('Micucci, Matthew P.', '', 'Pilots', 'Flexbase'),
    ('Moroz, Christopher R.', '', 'Pilots', 'Flexbase'),
    ('Pieton, Ryan S.', '', 'Pilots', 'Flexbase'),
    ('Silva, Ryan I.', '', 'Pilots', 'Flexbase'),
    ('Holmes, Thomas Y.', '', 'Pilots', 'Flexbase'),
    ('Howard, Analise', '', 'Pilots', 'Flexbase'),
    ('Reen, Gavin G.', '', 'Pilots', 'Flexbase'),
    ('McGinty, Garrett T.', '', 'Pilots', 'Flexbase'),
    ('Wienke, Moritz A.', 'Torq', 'Pilots', 'Flexbase'),
    ('Kolmer, Benjamin R.', '', 'Pilots', 'Flexbase'),
    ('Quimby, James R.', '', 'Pilots', 'Flexbase'),
    ('Regier, Kitt A.', 'Splash', 'Pilots', 'Flexbase'),
    ('Moore, Jaden R.', 'DI Moore', 'Pilots', 'Flexbase'),
    ('Hlavinka, Nathan C.', '', 'Pilots', 'Flexbase'),
    ('Mueser, William R.', '', 'Pilots', 'Flexbase'),
    ('Ponder, Stephen F.', 'DI Ponder', 'Pilots', 'Flexbase'),
    ('Woodall, Ryan F.', '', 'Pilots', 'Flexbase'),
    ('Evans, Noah R.', '', 'Pilots', 'Flexbase'),
    ('Samarah, Laith K.', 'DI Samarah', 'Pilots', 'Flexbase'),
    ('Thompson, Ben D.', '', 'Pilots', 'Flexbase'),
    ('Ortega, Soleil', '', 'AFE', 'Kysthotell'),
    ('Huckabone, Jake', '', 'AFE', 'Kysthotell'),
    ('Stennes, Trinidy', '', 'AFE', 'Kysthotell'),
    ('Prudente, Nicolas Mario', '', 'AFE', 'Kysthotell'),
    ('Hanlin, Stephanie', '', 'SARM', 'Kysthotell'),
    ('Lopez, Eduardo', '', 'SARM', 'Kysthotell'),
    ('Mahmoud, Yosef', '', 'SARM', 'Kysthotell'),
    ('Appleton, Audrey', '', 'SARM', 'Kysthotell'),
    ('Cunningham-Wray, Kassidy', '', 'Intel', 'Kysthotell'),
    ('Banks, Payton', '', 'Intel', 'Kysthotell'),
    ('Miller, Henry', '', 'Intel', 'Kysthotell'),
    ('Meadows, Kristen', '', 'SEL', 'Flexbase'),
    ('Anys, Michael', '', 'WX', 'Fosen Fjord'),
    ('Bratton, Joshua', '', 'WX', 'Fosen Fjord'),
    ('Watson, Austin', '', 'IDMT', 'Kysthotell'),
    ('Williams, Sean', '', 'ISSO/GSSO', 'Flexbase'),
    ('Farrington, Tony', '', 'ISSO/GSSO', 'Flexbase'),
    ('Andrews, Tim', '', 'OMS', 'Kysthotell'),
    ('Taylor, Nicholas', '', 'OMS', 'Kysthotell'),
    ('Ciambor, Dan', '', 'ALIS', 'Kysthotell'),
    ('Robison, Amber', '', 'ALIS', 'Kysthotell'),
    ('Mckenna, Aedan', '', 'ALIS', 'Kysthotell'),
    ('Sinn, Tyler', '', 'ALIS', 'Kysthotell'),
    ('Ciullo, Samuel', '', 'FSE', 'Kysthotell'),
]

cars = [
    ('Car 1',  'Skadi',    'EP69777',  'Yes', 5, '',           'Available'),
    ('Car 2',  'Thor',     'EN71155',  'No',  5, '',           'Available'),
    ('Car 3',  'Freya',    'EN35700',  'No',  5, '',           'Available'),
    ('Car 4',  'Tyr',      'EN46650',  'No',  5, '',           'Available'),
    ('Car 5',  'Loki',     'NH17739',  'No',  5, '',           'Available'),
    ('Car 6',  'Heimdall', 'AS68952',  'No',  5, '',           'Available'),
    ('Car 7',  'Baldur',   'EN43555',  'No',  5, '',           'Available'),
    ('Car 8',  'Vidar',    'EN53375',  'No',  5, '',           'Available'),
    ('Car 9',  'Bragi',    'EN87446',  'No',  5, '',           'Available'),
    ('Car 10', 'Fenrir',   'EN10871',  'No',  5, '',           'Available'),
    ('Car 11', 'Ragnar',   'OW75VLN',  'No',  5, '',           'Available'),
    ('Car 12', 'Odin',     'VJ36980',  'No',  5, 'Skooby',    'Permanent'),
    ('Car 13', 'Mjolnir',  'EP82941',  'Yes', 5, 'Tony/Sean', 'Permanent'),
]

car_names_assignable = ['Skadi','Thor','Freya','Tyr','Loki','Heimdall','Baldur','Vidar','Bragi','Fenrir','Ragnar']
car_names_all = [c[1] for c in cars]

# ============================================================
# TAB 1: CONFIG
# ============================================================
ws = wb.active
ws.title = 'Config'
ws.sheet_properties.tabColor = '2563EB'

ws.merge_cells('A1:H1')
ws['A1'] = 'VCO CAR PLAN - CONFIGURATION'
ws['A1'].font = header_font
ws['A1'].fill = blue_fill
ws['A1'].alignment = center

for label_cell, label, val_cell, val in [
    ('A3', 'TDY Location', 'B3', 'Orland, Norway'),
    ('A4', 'Start Date', 'B4', '2026-03-02'),
    ('A5', 'End Date', 'B5', '2026-03-24'),
    ('A6', 'VCO Name', 'B6', ''),
    ('A7', 'VCO Phone', 'B7', ''),
]:
    ws[label_cell] = label
    ws[label_cell].font = label_font
    ws[val_cell] = val

# Car inventory
ws.merge_cells('A9:H9')
ws['A9'] = 'CAR INVENTORY'
ws['A9'].font = subheader_font
ws['A9'].fill = dark_fill

for i, h in enumerate(['Car #', 'Norse Name', 'Plate', 'Electric', 'Capacity', 'Permanent Assign', 'Status', 'Notes'], 1):
    cell = ws.cell(row=10, column=i, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

for r, car in enumerate(cars, 11):
    for c, val in enumerate(car, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center
    if car[1] in ('Odin', 'Mjolnir'):
        ws.cell(row=r, column=8, value='Always assigned').border = thin_border

# Personnel roster
ws.merge_cells('A25:H25')
ws['A25'] = 'PERSONNEL ROSTER'
ws['A25'].font = subheader_font
ws['A25'].fill = dark_fill

for i, h in enumerate(['#', 'Name', 'Callsign', 'Section', 'Hotel', 'Fueler Auth', 'On-Site Start', 'On-Site End'], 1):
    cell = ws.cell(row=26, column=i, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

for idx, (name, callsign, section, hotel) in enumerate(personnel, 1):
    row = 26 + idx
    is_fueler = any(f in name for f in fuelers)
    ws.cell(row=row, column=1, value=idx).border = thin_border
    ws.cell(row=row, column=2, value=name).border = thin_border
    ws.cell(row=row, column=3, value=callsign).border = thin_border
    ws.cell(row=row, column=4, value=section).border = thin_border
    h_cell = ws.cell(row=row, column=5, value=hotel)
    h_cell.border = thin_border
    h_cell.fill = hotel_fills.get(hotel, PatternFill())
    ws.cell(row=row, column=6, value='Yes' if is_fueler else 'No').border = thin_border
    ws.cell(row=row, column=7, value='2026-03-02').border = thin_border
    ws.cell(row=row, column=8, value='2026-03-24').border = thin_border

for col, w in [(1,8),(2,28),(3,16),(4,12),(5,14),(6,12),(7,14),(8,14)]:
    ws.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# TAB 2: DAILY SCHEDULE
# ============================================================
ws2 = wb.create_sheet('Daily Schedule')
ws2.sheet_properties.tabColor = '10B981'

ws2.merge_cells('A1:M1')
ws2['A1'] = 'DAILY CAR ASSIGNMENTS'
ws2['A1'].font = header_font
ws2['A1'].fill = PatternFill('solid', fgColor='10B981')
ws2['A1'].alignment = center

ws2['A2'] = 'Date:'
ws2['A2'].font = label_font
ws2['A3'] = 'Show Time:'
ws2['A3'].font = label_font
ws2['B3'] = '0500'

sched_headers = ['Time Block', 'Event/Activity'] + car_names_assignable
for i, h in enumerate(sched_headers, 1):
    cell = ws2.cell(row=5, column=i, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

time_blocks = [
    ('0500-0530', 'Early Show / Step'),
    ('0530-0600', 'Brief'),
    ('0600-0700', 'Transit to Ops/Flightline'),
    ('0700-0800', 'Vul 1 Support'),
    ('0800-0900', 'Vul 1 Recovery'),
    ('0900-1000', 'Mid-Morning Turns'),
    ('1000-1100', 'Vul 2 Step/Brief'),
    ('1100-1200', 'Vul 2 Support'),
    ('1200-1300', 'Lunch Run'),
    ('1300-1400', 'Vul 2 Recovery / Debrief'),
    ('1400-1500', 'Afternoon'),
    ('1500-1600', 'Next Day Planning'),
    ('1600-1700', 'End of Duty Day'),
    ('1700-1800', 'Dinner Run'),
    ('1800-2000', 'Evening'),
    ('2000-2200', 'Late Evening'),
]

for r, (time, event) in enumerate(time_blocks, 6):
    ws2.cell(row=r, column=1, value=time).border = thin_border
    ws2.cell(row=r, column=2, value=event).border = thin_border
    for c in range(3, 14):
        ws2.cell(row=r, column=c).border = thin_border

# Passenger manifest
ws2.merge_cells('A24:M24')
ws2['A24'] = 'PASSENGER MANIFEST'
ws2['A24'].font = subheader_font
ws2['A24'].fill = dark_fill

manifest_headers = ['Time Block', 'Hotel Pickup'] + [f'{n} Pax' for n in car_names_assignable]
for i, h in enumerate(manifest_headers, 1):
    cell = ws2.cell(row=25, column=i, value=h)
    cell.font = Font(bold=True, size=9, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

for col, w in [(1,14),(2,26)] + [(i,14) for i in range(3,14)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# TAB 3: AVAILABILITY
# ============================================================
ws3 = wb.create_sheet('Availability')
ws3.sheet_properties.tabColor = 'F59E0B'

ws3.merge_cells('A1:G1')
ws3['A1'] = 'CAR AVAILABILITY - REAL-TIME STATUS'
ws3['A1'].font = header_font
ws3['A1'].fill = PatternFill('solid', fgColor='F59E0B')
ws3['A1'].alignment = center

ws3['A2'] = 'Date:'
ws3['A2'].font = label_font

for i, h in enumerate(['Car', 'Norse Name', 'Status', 'Driver', 'Destination', 'ETA Back', 'Notes'], 1):
    cell = ws3.cell(row=4, column=i, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

for r, (car_num, name, plate, elec, cap, perm, status) in enumerate(cars, 5):
    ws3.cell(row=r, column=1, value=car_num).border = thin_border
    ws3.cell(row=r, column=2, value=name).border = thin_border
    ws3.cell(row=r, column=3, value=status).border = thin_border
    if perm:
        ws3.cell(row=r, column=4, value=perm).border = thin_border
    for c in range(4 if not perm else 5, 8):
        ws3.cell(row=r, column=c).border = thin_border

# Sign-out log
ws3.merge_cells('A19:G19')
ws3['A19'] = 'CAR SIGN-OUT LOG'
ws3['A19'].font = subheader_font
ws3['A19'].fill = dark_fill

for i, h in enumerate(['Time Out', 'Car (Norse Name)', 'Driver', 'Passengers', 'Destination', 'ETA Back', 'Time In'], 1):
    cell = ws3.cell(row=20, column=i, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

for col, w in [(1,12),(2,16),(3,14),(4,24),(5,20),(6,12),(7,12)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# TAB 4: ROSTER
# ============================================================
ws4 = wb.create_sheet('Roster')
ws4.sheet_properties.tabColor = '8B5CF6'

ws4.merge_cells('A1:F1')
ws4['A1'] = 'DAILY ROSTER - WHO IS ON-SITE'
ws4['A1'].font = header_font
ws4['A1'].fill = PatternFill('solid', fgColor='8B5CF6')
ws4['A1'].alignment = center

ws4['A2'] = 'Date:'
ws4['A2'].font = label_font

for i, h in enumerate(['Name', 'Section', 'Hotel', 'Present', 'Override Notes', 'Fueler'], 1):
    cell = ws4.cell(row=4, column=i, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='374151')
    cell.alignment = center
    cell.border = thin_border

for idx, (name, callsign, section, hotel) in enumerate(personnel, 1):
    row = 4 + idx
    is_fueler = any(f in name for f in fuelers)
    ws4.cell(row=row, column=1, value=name).border = thin_border
    ws4.cell(row=row, column=2, value=section).border = thin_border
    h_cell = ws4.cell(row=row, column=3, value=hotel)
    h_cell.border = thin_border
    h_cell.fill = hotel_fills.get(hotel, PatternFill())
    ws4.cell(row=row, column=4, value='Yes').border = thin_border
    ws4.cell(row=row, column=5, value='').border = thin_border
    ws4.cell(row=row, column=6, value='Yes' if is_fueler else 'No').border = thin_border

for col, w in [(1,28),(2,12),(3,16),(4,10),(5,20),(6,10)]:
    ws4.column_dimensions[get_column_letter(col)].width = w

wb.save('VCO_Car_Plan.xlsx')
print(f'Built VCO_Car_Plan.xlsx: {len(personnel)} personnel, {len(cars)} cars, 4 tabs')
